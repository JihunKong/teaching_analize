"""
Excel Report Exporter
Generates comprehensive Excel workbooks with multiple sheets
"""

import logging
from typing import Dict, Any, List, Optional
from datetime import datetime
import io

# Excel generation
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, RadarChart, Reference
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExcelReportExporter:
    """Export analysis results to Excel format with professional formatting"""

    def __init__(self):
        """Initialize Excel exporter with style definitions"""
        # Color scheme
        self.colors = {
            'header_bg': 'FF667eea',
            'subheader_bg': 'FFE8EAF6',
            'success': 'FF4BC0C0',
            'warning': 'FFFFCE56',
            'danger': 'FFFF6384',
            'info': 'FF36A2EB',
            'light_gray': 'FFF8F9FA'
        }

        # Font styles
        self.fonts = {
            'header': Font(name='Arial', size=14, bold=True, color='FFFFFF'),
            'subheader': Font(name='Arial', size=12, bold=True, color='FF2C3E50'),
            'normal': Font(name='Arial', size=10),
            'bold': Font(name='Arial', size=10, bold=True)
        }

        # Border styles
        thin_border = Side(style='thin', color='FF000000')
        self.border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

    def export_to_excel(self, analysis_data: Dict[str, Any]) -> bytes:
        """
        Export analysis to Excel workbook

        Args:
            analysis_data: Complete analysis result

        Returns:
            Excel file as bytes
        """
        try:
            logger.info(f"Exporting to Excel: {analysis_data.get('analysis_id', 'unknown')}")

            framework = analysis_data.get('framework', 'generic')
            wb = Workbook()

            # Remove default sheet
            wb.remove(wb.active)

            # Create sheets based on framework
            if framework == 'cbil_comprehensive':
                self._create_cbil_comprehensive_sheets(wb, analysis_data)
            elif framework == 'cbil':
                self._create_cbil_sheets(wb, analysis_data)
            else:
                self._create_generic_sheets(wb, analysis_data)

            # Save to bytes
            excel_bytes = io.BytesIO()
            wb.save(excel_bytes)
            excel_bytes.seek(0)

            logger.info("Excel export completed successfully")
            return excel_bytes.getvalue()

        except Exception as e:
            logger.error(f"Excel export failed: {str(e)}")
            raise

    def _create_cbil_comprehensive_sheets(self, wb: Workbook, data: Dict[str, Any]):
        """Create sheets for CBIL comprehensive analysis"""
        result_data = data.get('result', data)

        # Sheet 1: Executive Summary
        self._create_executive_summary_sheet(wb, result_data)

        # Sheet 2: CBIL Scores
        self._create_cbil_scores_sheet(wb, result_data)

        # Sheet 3: Module 3 Metrics
        self._create_module3_metrics_sheet(wb, result_data)

        # Sheet 4: 3D Matrix
        self._create_3d_matrix_sheet(wb, result_data)

        # Sheet 5: Pattern Matching
        self._create_pattern_matching_sheet(wb, result_data)

        # Sheet 6: Coaching Feedback
        self._create_coaching_sheet(wb, result_data)

    def _create_executive_summary_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """Sheet 1: Executive Summary"""
        ws = wb.create_sheet("📊 Executive Summary")

        # Title
        ws.merge_cells('A1:E1')
        title_cell = ws['A1']
        title_cell.value = "CBIL + Module 3 종합 분석 보고서"
        title_cell.font = self.fonts['header']
        title_cell.fill = PatternFill(start_color=self.colors['header_bg'], end_color=self.colors['header_bg'], fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        # Metadata
        metadata = data.get('input_metadata', {})
        created_at = data.get('created_at', datetime.now().isoformat())

        row = 3
        info_items = [
            ('분석 ID:', data.get('evaluation_id', 'N/A')),
            ('생성 일시:', created_at.split('T')[0] if 'T' in created_at else created_at),
            ('총 발화 수:', metadata.get('total_utterances', 0)),
            ('CBIL 총점:', f"{metadata.get('cbil_total_score', 0)}/{metadata.get('cbil_max_score', 21)}"),
            ('CBIL 백분율:', f"{metadata.get('cbil_percentage', 0):.1f}%"),
            ('처리 시간:', f"{data.get('processing_time', 0):.2f}초")
        ]

        for label, value in info_items:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = self.fonts['bold']
            ws[f'B{row}'] = value
            row += 1

        # Pattern Matching
        row += 1
        ws[f'A{row}'] = "패턴 매칭 결과"
        ws[f'A{row}'].font = self.fonts['subheader']
        ws[f'A{row}'].fill = PatternFill(start_color=self.colors['subheader_bg'], end_color=self.colors['subheader_bg'], fill_type='solid')
        row += 1

        pattern_match = data.get('pattern_matching', {}).get('best_match', {})
        pattern_items = [
            ('최적 패턴:', pattern_match.get('pattern_name', 'N/A')),
            ('유사도:', f"{pattern_match.get('similarity_score', 0)*100:.1f}%"),
            ('매칭 품질:', pattern_match.get('match_quality', 'N/A'))
        ]

        for label, value in pattern_items:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = self.fonts['bold']
            ws[f'B{row}'] = value
            row += 1

        # CBIL Alignment
        cbil_alignment = data.get('pattern_matching', {}).get('cbil_alignment', 0)
        row += 1
        ws[f'A{row}'] = "CBIL-Module3 정렬 점수"
        ws[f'A{row}'].font = self.fonts['subheader']
        ws[f'A{row}'].fill = PatternFill(start_color=self.colors['subheader_bg'], end_color=self.colors['subheader_bg'], fill_type='solid')
        row += 1
        ws[f'A{row}'] = "정렬 점수:"
        ws[f'A{row}'].font = self.fonts['bold']
        ws[f'B{row}'] = f"{cbil_alignment*100:.1f}%"

        # Auto-size columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 35

    def _create_cbil_scores_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """Sheet 2: CBIL 7-Stage Scores"""
        ws = wb.create_sheet("🎯 CBIL Scores")

        # Title
        ws.merge_cells('A1:E1')
        title_cell = ws['A1']
        title_cell.value = "CBIL 7단계 점수"
        title_cell.font = self.fonts['header']
        title_cell.fill = PatternFill(start_color=self.colors['header_bg'], end_color=self.colors['header_bg'], fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 25

        # Headers
        headers = ['단계', '영문명', '점수', '최대점수', '백분율']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = self.fonts['bold']
            cell.fill = PatternFill(start_color=self.colors['subheader_bg'], end_color=self.colors['subheader_bg'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border

        # Extract CBIL scores
        cbil_insights = data.get('coaching_feedback', {}).get('cbil_insights', {})
        cbil_scores = cbil_insights.get('cbil_scores', {})
        stage_scores = cbil_scores.get('stage_scores', {})

        stage_info = [
            ('engage', '1. Engage', '흥미 유도 및 연결'),
            ('focus', '2. Focus', '탐구 방향 설정'),
            ('investigate', '3. Investigate', '자료 탐색 및 개념 형성'),
            ('organize', '4. Organize', '증거 조직화'),
            ('generalize', '5. Generalize', '일반화'),
            ('transfer', '6. Transfer', '전이'),
            ('reflect', '7. Reflect', '성찰')
        ]

        row = 4
        for stage_key, stage_num, stage_kr in stage_info:
            stage_data = stage_scores.get(stage_key, {})
            score = stage_data.get('score', 0)
            max_score = stage_data.get('max_score', 3)
            percentage = stage_data.get('percentage', 0)

            ws.cell(row=row, column=1).value = stage_num
            ws.cell(row=row, column=2).value = stage_kr
            ws.cell(row=row, column=3).value = score
            ws.cell(row=row, column=4).value = max_score
            ws.cell(row=row, column=5).value = f"{percentage:.1f}%"

            # Apply conditional formatting
            score_cell = ws.cell(row=row, column=3)
            if score >= 2.5:
                score_cell.fill = PatternFill(start_color=self.colors['success'], end_color=self.colors['success'], fill_type='solid')
            elif score >= 1.5:
                score_cell.fill = PatternFill(start_color=self.colors['warning'], end_color=self.colors['warning'], fill_type='solid')
            else:
                score_cell.fill = PatternFill(start_color=self.colors['danger'], end_color=self.colors['danger'], fill_type='solid')

            for col in range(1, 6):
                ws.cell(row=row, column=col).border = self.border
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')

            row += 1

        # Total
        row += 1
        total_score = cbil_scores.get('total_score', 0)
        max_total = cbil_scores.get('max_total_score', 21)
        overall_pct = cbil_scores.get('overall_percentage', 0)

        ws.merge_cells(f'A{row}:B{row}')
        total_cell = ws[f'A{row}']
        total_cell.value = "총점"
        total_cell.font = self.fonts['bold']
        total_cell.fill = PatternFill(start_color=self.colors['subheader_bg'], end_color=self.colors['subheader_bg'], fill_type='solid')
        total_cell.alignment = Alignment(horizontal='center')

        ws.cell(row=row, column=3).value = total_score
        ws.cell(row=row, column=3).font = self.fonts['bold']
        ws.cell(row=row, column=4).value = max_total
        ws.cell(row=row, column=4).font = self.fonts['bold']
        ws.cell(row=row, column=5).value = f"{overall_pct:.1f}%"
        ws.cell(row=row, column=5).font = self.fonts['bold']

        for col in range(1, 6):
            ws.cell(row=row, column=col).border = self.border

        # Auto-size
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def _create_module3_metrics_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """Sheet 3: Module 3 Quantitative Metrics"""
        ws = wb.create_sheet("📈 Module 3 Metrics")

        # Title
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = "Module 3 정량 지표 (15개)"
        title_cell.font = self.fonts['header']
        title_cell.fill = PatternFill(start_color=self.colors['header_bg'], end_color=self.colors['header_bg'], fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 25

        # Headers
        headers = ['지표명', '값', '정규화 점수', '최적 범위', '상태', '설명']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = self.fonts['bold']
            cell.fill = PatternFill(start_color=self.colors['subheader_bg'], end_color=self.colors['subheader_bg'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border

        # Data
        metrics = data.get('quantitative_metrics', {})
        row = 4

        for metric_name, metric_data in metrics.items():
            # Clean metric name
            display_name = metric_name.replace('_', ' ').title()

            ws.cell(row=row, column=1).value = display_name
            ws.cell(row=row, column=2).value = f"{metric_data.get('value', 0):.3f}"
            ws.cell(row=row, column=3).value = metric_data.get('normalized_score', 0)
            ws.cell(row=row, column=4).value = str(metric_data.get('optimal_range', 'N/A'))
            ws.cell(row=row, column=5).value = metric_data.get('status', 'N/A')
            ws.cell(row=row, column=6).value = metric_data.get('description', '')

            # Conditional formatting for status
            status_cell = ws.cell(row=row, column=5)
            status = metric_data.get('status', 'unknown')
            if status == 'optimal':
                status_cell.fill = PatternFill(start_color=self.colors['success'], end_color=self.colors['success'], fill_type='solid')
            elif status == 'good':
                status_cell.fill = PatternFill(start_color=self.colors['info'], end_color=self.colors['info'], fill_type='solid')
            elif status == 'needs_improvement':
                status_cell.fill = PatternFill(start_color=self.colors['warning'], end_color=self.colors['warning'], fill_type='solid')

            for col in range(1, 7):
                ws.cell(row=row, column=col).border = self.border

            row += 1

        # Auto-size
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 50

    def _create_3d_matrix_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """Sheet 4: 3D Matrix Data"""
        ws = wb.create_sheet("🎲 3D Matrix")

        # Title
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = "3D 매트릭스 (Stage × Context × Level)"
        title_cell.font = self.fonts['header']
        title_cell.fill = PatternFill(start_color=self.colors['header_bg'], end_color=self.colors['header_bg'], fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 25

        # Headers
        headers = ['Stage', 'Context', 'Level', '빈도', '백분율']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = self.fonts['bold']
            cell.fill = PatternFill(start_color=self.colors['subheader_bg'], end_color=self.colors['subheader_bg'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border

        # Extract matrix
        matrix_analysis = data.get('matrix_analysis', {})
        matrix = matrix_analysis.get('matrix', {})
        total_count = matrix_analysis.get('statistics', {}).get('total_utterances', 1)

        row = 4
        for stage_name, stage_data in matrix.items():
            for context_name, context_data in stage_data.items():
                for level_name, count in context_data.items():
                    if count > 0:
                        percentage = (count / total_count) * 100 if total_count > 0 else 0

                        ws.cell(row=row, column=1).value = stage_name.title()
                        ws.cell(row=row, column=2).value = context_name.title()
                        ws.cell(row=row, column=3).value = level_name.upper()
                        ws.cell(row=row, column=4).value = count
                        ws.cell(row=row, column=5).value = f"{percentage:.2f}%"

                        for col in range(1, 6):
                            ws.cell(row=row, column=col).border = self.border
                            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')

                        row += 1

        # Auto-size
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 18

    def _create_pattern_matching_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """Sheet 5: Pattern Matching"""
        ws = wb.create_sheet("🎯 Pattern Matching")

        # Title
        ws.merge_cells('A1:D1')
        title_cell = ws['A1']
        title_cell.value = "패턴 매칭 분석"
        title_cell.font = self.fonts['header']
        title_cell.fill = PatternFill(start_color=self.colors['header_bg'], end_color=self.colors['header_bg'], fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 25

        # Best Match
        pattern_match = data.get('pattern_matching', {}).get('best_match', {})
        row = 3

        ws[f'A{row}'] = "최적 패턴:"
        ws[f'A{row}'].font = self.fonts['bold']
        ws[f'B{row}'] = pattern_match.get('pattern_name', 'N/A')
        row += 1

        ws[f'A{row}'] = "유사도:"
        ws[f'A{row}'].font = self.fonts['bold']
        ws[f'B{row}'] = f"{pattern_match.get('similarity_score', 0)*100:.1f}%"
        row += 1

        ws[f'A{row}'] = "매칭 품질:"
        ws[f'A{row}'].font = self.fonts['bold']
        ws[f'B{row}'] = pattern_match.get('match_quality', 'N/A')
        row += 2

        # Stage Similarities
        ws[f'A{row}'] = "단계별 유사도"
        ws[f'A{row}'].font = self.fonts['subheader']
        ws[f'A{row}'].fill = PatternFill(start_color=self.colors['subheader_bg'], end_color=self.colors['subheader_bg'], fill_type='solid')
        row += 1

        stage_similarities = pattern_match.get('stage_similarities', {})
        for stage, similarity in stage_similarities.items():
            ws[f'A{row}'] = stage.title()
            ws[f'B{row}'] = f"{similarity*100:.1f}%"
            row += 1

        # Auto-size
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30

    def _create_coaching_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """Sheet 6: Coaching Feedback"""
        ws = wb.create_sheet("💡 Coaching")

        # Title
        ws.merge_cells('A1:C1')
        title_cell = ws['A1']
        title_cell.value = "코칭 피드백"
        title_cell.font = self.fonts['header']
        title_cell.fill = PatternFill(start_color=self.colors['header_bg'], end_color=self.colors['header_bg'], fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 25

        coaching = data.get('coaching_feedback', {})
        row = 3

        # Overall Assessment
        ws[f'A{row}'] = "전체 평가"
        ws[f'A{row}'].font = self.fonts['subheader']
        ws[f'A{row}'].fill = PatternFill(start_color=self.colors['subheader_bg'], end_color=self.colors['subheader_bg'], fill_type='solid')
        row += 1
        ws.merge_cells(f'A{row}:C{row+2}')
        ws[f'A{row}'] = coaching.get('overall_assessment', '')
        ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        row += 4

        # Strengths
        ws[f'A{row}'] = "강점"
        ws[f'A{row}'].font = self.fonts['subheader']
        ws[f'A{row}'].fill = PatternFill(start_color=self.colors['subheader_bg'], end_color=self.colors['subheader_bg'], fill_type='solid')
        row += 1
        for strength in coaching.get('strengths', []):
            ws.merge_cells(f'A{row}:C{row}')
            ws[f'A{row}'] = f"• {strength}"
            ws[f'A{row}'].alignment = Alignment(wrap_text=True)
            row += 1

        row += 1

        # Priority Actions
        ws[f'A{row}'] = "우선 조치"
        ws[f'A{row}'].font = self.fonts['subheader']
        ws[f'A{row}'].fill = PatternFill(start_color=self.colors['subheader_bg'], end_color=self.colors['subheader_bg'], fill_type='solid')
        row += 1
        for action in coaching.get('priority_actions', []):
            ws.merge_cells(f'A{row}:C{row}')
            ws[f'A{row}'] = f"• {action}"
            ws[f'A{row}'].alignment = Alignment(wrap_text=True)
            row += 1

        # Auto-size
        ws.column_dimensions['A'].width = 100

    def _create_cbil_sheets(self, wb: Workbook, data: Dict[str, Any]):
        """Create sheets for CBIL-only analysis"""
        # Simplified version for CBIL only
        self._create_cbil_scores_sheet(wb, {'coaching_feedback': {'cbil_insights': {'cbil_scores': {}}}})

    def _create_generic_sheets(self, wb: Workbook, data: Dict[str, Any]):
        """Create sheets for generic analysis"""
        ws = wb.create_sheet("분석 결과")
        ws['A1'] = "분석 ID:"
        ws['B1'] = data.get('analysis_id', 'N/A')
        ws['A2'] = "프레임워크:"
        ws['B2'] = data.get('framework', 'N/A')
        ws['A3'] = "생성 시간:"
        ws['B3'] = data.get('created_at', datetime.now().isoformat())


# ============ Demo/Test Code ============

if __name__ == "__main__":
    print("=" * 60)
    print("Excel Report Exporter Demo")
    print("=" * 60)
    print("\nThis module exports analysis results to Excel format.")
    print("Usage: Import and use in main.py endpoints")
    print("\nFeatures:")
    print("  ✓ Multi-sheet workbooks")
    print("  ✓ Professional formatting")
    print("  ✓ Conditional formatting")
    print("  ✓ Auto-sizing columns")
    print("  ✓ Border styling")
    print("\n" + "=" * 60)
